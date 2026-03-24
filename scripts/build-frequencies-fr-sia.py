from pathlib import Path
import json
import re
import unicodedata

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_XLSX = BASE_DIR / "data-source" / "sia" / "_AIP_XML_SIA_AAAA-MM-JJ.xlsx"
OUTPUT_JSON = BASE_DIR / "public" / "data" / "frequencies-fr-sia.json"


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_space(value):
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def normalize_code(value):
    return normalize_space(value).upper()


def fold_text(value):
    text = normalize_space(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.lower()


def to_float_or_none(value):
    text = normalize_space(value).replace(",", ".")
    if not text:
        return None

    try:
        parsed = float(text)
    except Exception:
        return None

    if parsed <= 0:
        return None

    return parsed


def sheet_to_dicts(workbook, sheet_name):
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_space(h) for h in rows[0]]

    records = []
    for row in rows[1:]:
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        records.append(record)

    return records


def iter_link_values(row):
    for key, value in row.items():
        key_norm = normalize_space(key).lower()
        if key_norm.startswith("lk"):
            text = normalize_space(value)
            if text:
                yield text


def extract_bracket_airport_ident(text):
    text = normalize_space(text)
    match = re.search(r"\[([A-Z0-9]{2})\]\[([A-Z0-9]{2})\]", text)
    if match:
        return f"{normalize_code(match.group(1))}{normalize_code(match.group(2))}"
    return ""


def extract_four_letter_ident(text):
    text = normalize_code(text)
    match = re.search(r"\b([A-Z]{4})\b", text)
    if match:
        return match.group(1)
    return ""


def extract_two_letter_ad_code(text):
    text = normalize_space(text)
    match = re.search(r"\[[A-Z0-9]{2}\]\[([A-Z0-9]{2})\]", text)
    if match:
        return normalize_code(match.group(1))
    return ""


def first_non_empty(*values):
    for value in values:
        text = normalize_space(value)
        if text:
            return text
    return ""


def build_ad_index(ad_rows):
    index = {}

    for row in ad_rows:
        ad_code = normalize_code(
            first_non_empty(
                row.get("AdCode"),
                row.get("Code"),
                row.get("CodeAd"),
            )
        )
        if not ad_code:
            continue

        airport_ident = ""
        for value in row.values():
            airport_ident = (
                extract_bracket_airport_ident(value)
                or extract_four_letter_ident(value)
            )
            if airport_ident:
                break

        index[ad_code] = {
            "ad_code": ad_code,
            "airport_ident": airport_ident,
            "airport_name": first_non_empty(
                row.get("AdNomComplet"),
                row.get("AdNomCarto"),
                row.get("Nom"),
            ),
            "airport_status": first_non_empty(
                row.get("AdStatut"),
                row.get("Statut"),
            ),
            "airport_situation": first_non_empty(
                row.get("AdSituation"),
                row.get("Situation"),
            ),
        }

    return index


def extract_service_airport_ident(service_row, ad_index):
    for value in iter_link_values(service_row):
        ident = extract_bracket_airport_ident(value)
        if ident:
            return ident

    for candidate in [
        service_row.get("IndicLieu"),
        service_row.get("IndicService"),
        service_row.get("Service"),
        *service_row.values(),
    ]:
        ident = extract_four_letter_ident(candidate)
        if ident:
            return ident

    ad_code = normalize_code(
        first_non_empty(
            service_row.get("AdCode"),
            service_row.get("Code"),
            service_row.get("CodeAd"),
        )
    )

    if not ad_code:
        for value in iter_link_values(service_row):
            ad_code = extract_two_letter_ad_code(value)
            if ad_code:
                break

    if ad_code and ad_code in ad_index:
        return normalize_code(ad_index[ad_code].get("airport_ident"))

    return ""


def canonical_service_name_from_text(text):
    folded = fold_text(text)
    if not folded or folded == ".":
        return ""

    mapping = {
        "twr": "Tour",
        "tour": "Tour",
        "tower": "Tour",
        "gnd": "Sol",
        "ground": "Sol",
        "sol": "Sol",
        "app": "Approche",
        "approach": "Approche",
        "approche": "Approche",
        "atis": "ATIS",
        "d-atis": "D-ATIS",
        "datis": "D-ATIS",
        "info": "Info",
        "information": "Info",
        "fis": "Info",
        "afis": "AFIS",
        "a/a": "Auto-information",
        "aa": "Auto-information",
        "auto-information": "Auto-information",
        "auto information": "Auto-information",
        "unicom": "UNICOM",
        "ctaf": "CTAF",
        "prevol": "Prévol",
        "pre-flight": "Prévol",
        "preflight": "Prévol",
        "trafic": "Trafic",
        "traffic": "Trafic",
        "dep": "Départ",
        "depart": "Départ",
        "arr": "Arrivée",
        "arrivee": "Arrivée",
        "rad": "Radar",
        "radar": "Radar",
        "acc": "Contrôle",
        "acs": "Contrôle",
        "controle": "Contrôle",
        "control": "Contrôle",
    }

    return mapping.get(folded, "")


def resolve_service_name(service_code, indic_service):
    service_code = normalize_code(service_code)
    indic_service = normalize_space(indic_service)

    if indic_service == ".":
        indic_service = ""

    if service_code == "A/A":
        return "Auto-information", ""

    service_name_from_detail = canonical_service_name_from_text(indic_service)
    service_name_from_code = (
        canonical_service_name_from_text(service_code) or service_code or "Fréquence"
    )

    if service_name_from_detail:
        return service_name_from_detail, ""

    extra_detail = ""
    if indic_service:
        folded_detail = fold_text(indic_service)
        folded_base = fold_text(service_name_from_code)
        folded_code = fold_text(service_code)

        if folded_detail not in {folded_base, folded_code}:
            extra_detail = indic_service

    return service_name_from_code, extra_detail


def build_service_label(service_row):
    service_code = normalize_code(service_row.get("Service"))
    indic_lieu = normalize_space(service_row.get("IndicLieu"))
    indic_service = normalize_space(service_row.get("IndicService"))

    service_name, extra_detail = resolve_service_name(service_code, indic_service)

    parts = []

    if service_name:
        parts.append(service_name)

    if indic_lieu:
        folded_place = fold_text(indic_lieu)
        folded_service_name = fold_text(service_name)

        if folded_place != folded_service_name:
            parts.append(indic_lieu)

    if extra_detail:
        folded_extra = fold_text(extra_detail)
        folded_service_name = fold_text(service_name)
        folded_place = fold_text(indic_lieu)

        if folded_extra not in {folded_service_name, folded_place}:
            parts.append(extra_detail)

    label = " ".join(part for part in parts if part).strip()
    return label or service_name or service_code or "Fréquence"


def normalize_frequency_type(service_code, label=""):
    label_folded = fold_text(label)
    code = normalize_code(service_code)

    if label_folded in {"tour"} or code == "TWR":
        return "tower"

    if label_folded in {"sol"} or code in {"GND", "SOL"}:
        return "ground"

    if label_folded in {"approche"} or code == "APP":
        return "approach"

    if label_folded in {"info"} or code in {"FIS", "INFO"}:
        return "information"

    if label_folded == "afis" or code == "AFIS":
        return "information"

    if label_folded in {"atis", "d-atis"} or code == "ATIS":
        return "atis"

    if label_folded in {"auto-information", "unicom", "ctaf"} or code == "A/A":
        return "air_to_air"

    if label_folded == "prevol":
        return "preflight"

    if label_folded == "trafic":
        return "traffic"

    if label_folded == "depart" or code == "DEP":
        return "departure"

    if label_folded == "arrivee" or code == "ARR":
        return "arrival"

    if label_folded == "radar" or code == "RAD":
        return "radar"

    if label_folded == "controle" or code in {"ACC", "ACS"}:
        return "control"

    return "other"


def is_exploitable_frequency(entry):
    label = normalize_code(entry.get("label"))
    service_code = normalize_code(entry.get("service_code"))
    remark = normalize_code(entry.get("remark"))

    if entry.get("frequency_mhz") is None:
        return False

    if service_code == "VDF":
        return False

    if "GONIO" in label or "GONIO" in remark:
        return False

    if "HOMER" in label or "HOMER" in remark:
        return False

    return True


def frequency_priority(entry):
    order = {
        "tower": 0,
        "ground": 1,
        "information": 2,
        "air_to_air": 3,
        "approach": 4,
        "preflight": 5,
        "traffic": 6,
        "departure": 7,
        "arrival": 8,
        "atis": 9,
        "radar": 10,
        "control": 11,
        "other": 12,
    }

    return order.get(entry.get("frequency_type", "other"), 99)


def frequency_entry_score(entry):
    score = 0

    if normalize_space(entry.get("sector")):
        score += 20

    if normalize_space(entry.get("remark")):
        score += 10

    if normalize_space(entry.get("hours_text")):
        score += 4

    if normalize_space(entry.get("hours_code")):
        score += 2

    suppletive = fold_text(entry.get("suppletive"))
    if suppletive and suppletive != "non":
        score += 1

    return score


def dedupe_frequencies(entries):
    by_key = {}

    for entry in entries:
        key = (
            normalize_code(entry.get("frequency_type")),
            normalize_space(entry.get("label")).lower(),
            entry.get("frequency_mhz"),
            normalize_space(entry.get("sector")).lower(),
            normalize_space(entry.get("suppletive")).lower(),
        )

        current = by_key.get(key)
        if current is None or frequency_entry_score(entry) > frequency_entry_score(current):
            by_key[key] = entry

    return list(by_key.values())


def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Fichier source introuvable : {INPUT_XLSX}")

    workbook = load_workbook(INPUT_XLSX, read_only=True, data_only=True)

    required_sheets = ["FrequenceS", "ServiceS", "AdS"]
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille manquante dans le fichier SIA : {sheet_name}")

    frequency_rows = sheet_to_dicts(workbook, "FrequenceS")
    service_rows = sheet_to_dicts(workbook, "ServiceS")
    ad_rows = sheet_to_dicts(workbook, "AdS")

    ad_index = build_ad_index(ad_rows)

    service_by_pk = {}
    unresolved_services = 0

    for row in service_rows:
        pk = row.get("pk")
        if pk is None:
            continue

        try:
            pk = int(pk)
        except Exception:
            continue

        airport_ident = extract_service_airport_ident(row, ad_index)
        if not airport_ident:
            unresolved_services += 1
            continue

        service_code = normalize_code(row.get("Service"))
        service_label = build_service_label(row)
        indic_lieu = normalize_space(row.get("IndicLieu"))
        indic_service = normalize_space(row.get("IndicService"))
        language = normalize_space(row.get("Langue"))

        service_by_pk[pk] = {
            "pk": pk,
            "airport_ident": airport_ident,
            "service_code": service_code,
            "label": service_label,
            "indic_lieu": indic_lieu,
            "indic_service": indic_service,
            "language": language,
        }

    grouped = {}

    skipped_non_exploitable = 0

    for row in frequency_rows:
        service_pk = row.get("pk2")
        if service_pk is None:
            continue

        try:
            service_pk = int(service_pk)
        except Exception:
            continue

        service = service_by_pk.get(service_pk)
        if not service:
            continue

        frequency_mhz = to_float_or_none(row.get("Frequence"))
        if frequency_mhz is None:
            skipped_non_exploitable += 1
            continue

        airport_ident = service["airport_ident"]
        ad_code = airport_ident[2:] if len(airport_ident) == 4 else ""
        ad_info = ad_index.get(ad_code, {})

        frequency_entry = {
            "frequency_type": normalize_frequency_type(
                service["service_code"],
                service["label"],
            ),
            "service_code": service["service_code"],
            "label": service["label"],
            "frequency_mhz": round(frequency_mhz, 3),
            "hours_code": normalize_space(row.get("HorCode")),
            "hours_text": normalize_space(row.get("HorTxt")),
            "spacing": normalize_space(row.get("Espacement")),
            "remark": normalize_space(row.get("Remarque")),
            "sector": normalize_space(row.get("SecteurSituation")),
            "suppletive": normalize_space(row.get("Suppletive")),
            "language": service["language"],
            "indic_lieu": service["indic_lieu"],
            "indic_service": service["indic_service"],
        }

        if not is_exploitable_frequency(frequency_entry):
            skipped_non_exploitable += 1
            continue

        if airport_ident not in grouped:
            grouped[airport_ident] = {
                "airport_ident": airport_ident,
                "weather_code": airport_ident if len(airport_ident) == 4 else "",
                "airport_name": ad_info.get("airport_name", ""),
                "airport_status": ad_info.get("airport_status", ""),
                "airport_situation": ad_info.get("airport_situation", ""),
                "source": "SIA XLSX",
                "frequencies": [],
            }

        grouped[airport_ident]["frequencies"].append(frequency_entry)

    result = []

    for airport_ident, item in grouped.items():
        frequencies = dedupe_frequencies(item["frequencies"])
        frequencies.sort(
            key=lambda entry: (
                frequency_priority(entry),
                normalize_space(entry.get("label")).lower(),
                normalize_space(entry.get("sector")).lower(),
                entry.get("frequency_mhz", 9999),
            )
        )

        result.append(
            {
                **item,
                "frequencies": frequencies,
                "primary_frequency": frequencies[0] if frequencies else None,
            }
        )

    result.sort(key=lambda item: item["airport_ident"])

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Base fréquences SIA générée : {len(result)} terrains exportés")
    print(f"Services non rattachés à un terrain : {unresolved_services}")
    print(f"Fréquences non exploitables ignorées : {skipped_non_exploitable}")
    print(f"Fichier écrit : {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
